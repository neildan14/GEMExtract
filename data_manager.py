import pandas as pd
import sqlite3
from datetime import datetime
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def get_connection():
    """Get database connection."""
    return sqlite3.connect("data/app_database.db")

def import_excel(file_path):
    """
    Import data from an Excel file into the database.
    Maps columns by POSITION (column 0, 1, 2, etc) rather than by name.
    
    Args:
        file_path (str): Path to the Excel file to import
        
    Returns:
        dict: Contains 'success' (bool), 'message' (str), and 'rows_inserted' (int)
    """
    try:
        # Read Excel file
        df = pd.read_excel(file_path)
        
        if df.empty:
            return {
                'success': False,
                'message': "Excel file is empty",
                'rows_inserted': 0
            }
        
        # Store original column info for diagnostics
        original_columns = df.columns.tolist()
        num_rows = len(df)
        num_cols = len(df.columns)
        
        # Map columns by POSITION (not by name)
        # This works even with unnamed columns like "Unnamed: 0", "Unnamed: 1", etc.
        position_mapping = {
            0: 'NO',
            1: 'CLIENT',
            2: 'LOCATION',
            3: 'CURRENCY',
            4: 'PROGRESS',
            5: 'IMAGE_PATH',
            6: 'PROJECT',
            7: 'GEMPHIL_DEVICE',
            8: 'DETAILS',
            9: 'CONTACT_PERSON',
            10: 'EMAIL'
        }
        
        # Create new dataframe with only the columns we need (up to 11 columns)
        cols_to_use = min(11, num_cols)  # Use up to 11 columns (11 data columns)
        df_import = df.iloc[:, :cols_to_use].copy()
        
        # Rename columns by position
        new_col_names = {}
        for pos in range(cols_to_use):
            if pos in position_mapping:
                new_col_names[pos] = position_mapping[pos]
        
        df_import.columns = [new_col_names.get(i, f'UNUSED_{i}') for i in range(cols_to_use)]
        
        # Debug info
        debug_first_row = {}
        if not df_import.empty:
            first_row_data = df_import.iloc[0]
            for col in ['NO', 'CLIENT', 'PROJECT', 'GEMPHIL_DEVICE', 'LOCATION']:
                if col in first_row_data:
                    val = str(first_row_data[col]).strip() if pd.notna(first_row_data[col]) else 'EMPTY'
                    debug_first_row[col] = val
        
        # Connect to database
        conn = get_connection()
        cursor = conn.cursor()
        
        rows_inserted = 0
        errors = []
        
        # Insert each row
        for idx, row in df_import.iterrows():
            try:
                # Get values with fallbacks and NaN handling
                no_val = str(row.get('NO', '')).strip() if pd.notna(row.get('NO', '')) else ''
                client_val = str(row.get('CLIENT', '')).strip() if pd.notna(row.get('CLIENT', '')) else ''
                location_val = str(row.get('LOCATION', '')).strip() if pd.notna(row.get('LOCATION', '')) else ''
                currency_val = str(row.get('CURRENCY', '')).strip() if pd.notna(row.get('CURRENCY', '')) else ''
                progress_val = str(row.get('PROGRESS', '')).strip() if pd.notna(row.get('PROGRESS', '')) else ''
                image_path_val = str(row.get('IMAGE_PATH', '')).strip() if pd.notna(row.get('IMAGE_PATH', '')) else ''
                project_val = str(row.get('PROJECT', '')).strip() if pd.notna(row.get('PROJECT', '')) else ''
                gemphil_device_val = str(row.get('GEMPHIL_DEVICE', '')).strip() if pd.notna(row.get('GEMPHIL_DEVICE', '')) else ''
                details_val = str(row.get('DETAILS', '')).strip() if pd.notna(row.get('DETAILS', '')) else ''
                contact_person_val = str(row.get('CONTACT_PERSON', '')).strip() if pd.notna(row.get('CONTACT_PERSON', '')) else ''
                email_val = str(row.get('EMAIL', '')).strip() if pd.notna(row.get('EMAIL', '')) else ''
                
                cursor.execute("""
                    INSERT INTO records 
                    (NO, CLIENT, LOCATION, CURRENCY, PROGRESS, IMAGE_PATH, PROJECT, 
                     GEMPHIL_DEVICE, DETAILS, CONTACT_PERSON, EMAIL)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    no_val,
                    client_val,
                    location_val,
                    currency_val,
                    progress_val,
                    image_path_val,
                    project_val,
                    gemphil_device_val,
                    details_val,
                    contact_person_val,
                    email_val
                ))
                rows_inserted += 1
            except Exception as row_error:
                errors.append(f"Row {idx + 2}: {str(row_error)}")
        
        conn.commit()
        conn.close()
        
        if rows_inserted == 0:
            return {
                'success': False,
                'message': f"No rows were imported.\n\nFile has {num_rows} rows × {num_cols} columns\n\nCheck that your Excel file has data in the proper columns.",
                'rows_inserted': 0
            }
        
        # Verify data was actually imported by checking first row
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT CLIENT, PROJECT, GEMPHIL_DEVICE, NO, LOCATION FROM records ORDER BY id DESC LIMIT 1")
        first_row = cursor.fetchone()
        conn.close()
        
        message = f"✓ Successfully imported {rows_inserted} rows\n\n"
        message += f"📋 Column Mapping (Position → Database Field):\n"
        for pos in range(min(12, num_cols)):
            if pos in position_mapping:
                message += f"  • Column {pos} → {position_mapping[pos]}\n"
        
        message += f"\n📝 Sample Data (Excel Row 1):\n"
        for key, val in debug_first_row.items():
            message += f"  • {key}: {val}\n"
        
        if first_row:
            message += f"\n💾 Data In Database (After Insert):\n"
            message += f"  • NO: {first_row[3] if first_row[3] else 'EMPTY'}\n"
            message += f"  • CLIENT: {first_row[0] if first_row[0] else 'EMPTY'}\n"
            message += f"  • LOCATION: {first_row[4] if first_row[4] else 'EMPTY'}\n"
            message += f"  • PROJECT: {first_row[1] if first_row[1] else 'EMPTY'}\n"
            message += f"  • GEMPHIL_DEVICE: {first_row[2] if first_row[2] else 'EMPTY'}"
        
        if errors:
            message += f"\n\n⚠️ Errors:\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                message += f"\n... and {len(errors) - 5} more warnings"
        
        return {
            'success': True,
            'message': message,
            'rows_inserted': rows_inserted
        }
        
    except FileNotFoundError:
        return {
            'success': False,
            'message': f"File not found: {file_path}",
            'rows_inserted': 0
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"Error importing Excel file: {str(e)}",
            'rows_inserted': 0
        }

def export_excel(record_ids=None, selected_columns=None, save_path=None, preserve_order=False):
    """
    Export data from database to an Excel file.
    
    Args:
        record_ids (list): List of record IDs to export. If None, exports all records.
        selected_columns (list): List of column names to export. If None, exports all columns.
                                Options: ['NO', 'CLIENT', 'LOCATION', 'CURRENCY', 'PROGRESS',
                                         'IMAGE_PATH', 'PROJECT', 'GEMPHIL_DEVICE', 
                                         'DETAILS', 'CONTACT_PERSON', 'EMAIL']
        save_path (str): Path where the Excel file should be saved. 
                        If None, creates a timestamped file in the data folder.
        preserve_order (bool): If True and record_ids provided, preserves the order of record_ids
                              instead of ordering by id. Default False for backward compatibility.
        
    Returns:
        dict: Contains 'success' (bool), 'message' (str), and 'file_path' (str)
    """
    try:
        # Default columns to export (exclude id)
        default_columns = [
            'NO', 'CLIENT', 'LOCATION', 'CURRENCY', 'PROGRESS', 'IMAGE_PATH',
            'PROJECT', 'GEMPHIL_DEVICE', 'DETAILS', 'CONTACT_PERSON', 'EMAIL'
        ]
        
        # Use selected columns or all columns
        columns_to_export = selected_columns if selected_columns else default_columns
        
        # Validate columns
        invalid_columns = set(columns_to_export) - set(default_columns)
        if invalid_columns:
            return {
                'success': False,
                'message': f"Invalid columns: {', '.join(invalid_columns)}",
                'file_path': None
            }
        
        # Build SQL query
        column_names = ', '.join(columns_to_export)
        
        if record_ids:
            # Export specific records
            if preserve_order and len(record_ids) > 0:
                # Create CASE statement to preserve order of record_ids
                case_parts = [f"WHEN id = {rid} THEN {idx}" for idx, rid in enumerate(record_ids)]
                case_statement = "CASE " + " ".join(case_parts) + " ELSE 999999 END"
                placeholders = ','.join('?' * len(record_ids))
                query = f"SELECT {column_names} FROM records WHERE id IN ({placeholders}) ORDER BY {case_statement}"
                params = record_ids
            else:
                # Default behavior: order by id
                placeholders = ','.join('?' * len(record_ids))
                query = f"SELECT {column_names} FROM records WHERE id IN ({placeholders}) ORDER BY id ASC"
                params = record_ids
        else:
            # Export all records
            query = f"SELECT {column_names} FROM records ORDER BY id ASC"
            params = []
        
        # Fetch data from database
        conn = get_connection()
        if params:
            df = pd.read_sql_query(query, conn, params=params)
        else:
            df = pd.read_sql_query(query, conn)
        conn.close()
        
        # Generate default filename if not provided
        if not save_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            save_path = f"data/export_{timestamp}.xlsx"
        
        # Create Excel file with formatting
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Records')
            
            worksheet = writer.sheets['Records']
            
            # Apply formatting to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    # Enable text wrapping
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    # Optional: Make header bold
                    if cell.row == 1:
                        cell.font = Font(bold=True)
            
            # Set column widths and auto-adjust row heights
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Calculate max length for column width
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width (capped at 50 for readability)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Auto-adjust row heights based on content
            for row in worksheet.iter_rows():
                max_height = 15  # Default minimum height
                
                for cell in row:
                    if cell.value:
                        cell_text = str(cell.value)
                        # Estimate lines based on text length and column width
                        col_width = worksheet.column_dimensions[get_column_letter(cell.column)].width
                        # Account for text wrapping: estimate number of lines
                        estimated_lines = (len(cell_text) // max(int(col_width), 10)) + 1
                        # Calculate row height (approximately 15 pixels per line)
                        estimated_height = estimated_lines * 15
                        max_height = max(max_height, estimated_height)
                
                worksheet.row_dimensions[row[0].row].height = max_height
        
        message = f"Successfully exported {len(df)} records to {save_path}"
        
        return {
            'success': True,
            'message': message,
            'file_path': save_path
        }
        
    except Exception as e:
        return {
            'success': False,
            'message': f"Error exporting to Excel: {str(e)}",
            'file_path': None
        }

def get_column_options():
    """
    Get available columns for export selection.
    
    Returns:
        list: List of available column names
    """
    return [
        'NO', 'CLIENT', 'LOCATION', 'CURRENCY', 'PROGRESS', 'IMAGE_PATH',
        'PROJECT', 'GEMPHIL_DEVICE', 'DETAILS', 'CONTACT_PERSON', 'EMAIL'
    ]
